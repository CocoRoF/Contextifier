"""
Comprehensive tests for the XLS handler pipeline.

Tests cover:
1.  Constants
2.  XlsConverter: OLE2 validation, xlrd opening, close
3.  XlsPreprocessor: Unwrap, sheet info
4.  XlsMetadataExtractor: OLE + xlrd metadata
5.  Layout detection: LayoutRange, layout_detect_range, object_detect
6.  Table conversion: Markdown / HTML
7.  XlsContentExtractor: Sheet iteration, table output
8.  XLSHandler: Wiring, delegation
9.  Full pipeline
10. Edge cases
"""

from __future__ import annotations

import io
import struct
import zipfile
from datetime import datetime
from typing import Any, List
from unittest.mock import MagicMock, patch, PropertyMock

import xlrd
import pytest

# ═══════════════════════════════════════════════════════════════════════════════
# Imports under test
# ═══════════════════════════════════════════════════════════════════════════════

from contextifier_new.handlers.xls._constants import (
    OLE2_MAGIC,
    ZIP_MAGIC,
    MAX_SCAN_ROWS,
    MAX_SCAN_COLS,
)
from contextifier_new.handlers.xls.converter import (
    XlsConverter,
    XlsConvertedData,
)
from contextifier_new.handlers.xls.preprocessor import XlsPreprocessor
from contextifier_new.handlers.xls.metadata_extractor import XlsMetadataExtractor
from contextifier_new.handlers.xls._layout import (
    LayoutRange,
    layout_detect_range,
    object_detect,
)
from contextifier_new.handlers.xls._table import (
    convert_region_to_table,
    convert_region_to_markdown,
    convert_region_to_html,
    convert_sheet_to_text,
    _format_cell,
)
from contextifier_new.handlers.xls.content_extractor import XlsContentExtractor
from contextifier_new.handlers.xls.handler import XLSHandler
from contextifier_new.types import (
    DocumentMetadata,
    ExtractionResult,
    FileContext,
    PreprocessedData,
    TableData,
)
from contextifier_new.config import ProcessingConfig
from contextifier_new.errors import ConversionError, PreprocessingError


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _make_ctx(data: bytes, ext: str = "xls", name: str = "test.xls") -> FileContext:
    return {
        "file_data": data,
        "file_extension": ext,
        "file_name": name,
        "file_path": f"/tmp/{name}",
        "file_category": "spreadsheet",
        "file_stream": io.BytesIO(data),
        "file_size": len(data),
    }


def _make_xls_bytes() -> bytes:
    """Create a minimal XLS file using xlrd's test data. We write BIFF directly."""
    # Use xlrd to parse, but we need a real XLS file.
    # xlrd 2.x can only read .xls files (not write). We'll create a minimal
    # BIFF8 file by hand — or more practically, use a tiny known-good XLS.
    # For testing, we'll construct a mock OLE file that xlrd can read.
    # Actually, the simplest approach: create via openpyxl (XLSX) won't work.
    # Let's use a pragmatic approach with mock objects for unit tests.
    return None  # Marker: use mocks instead


def _mock_xlrd_book(
    sheets: dict[str, list[list[Any]]] | None = None,
    user_name: str = "",
    datemode: int = 0,
) -> MagicMock:
    """Create a mock xlrd.Book with specified sheet data."""
    sheets = sheets or {"Sheet1": [["H1", "H2"], ["V1", "V2"]]}

    mock_book = MagicMock()
    mock_book.nsheets = len(sheets)
    mock_book.user_name = user_name
    mock_book.datemode = datemode
    mock_book.xf_list = []
    mock_book.release_resources = MagicMock()

    sheet_list = []
    for name, rows in sheets.items():
        mock_sheet = _mock_xlrd_sheet(name, rows)
        sheet_list.append(mock_sheet)

    def sheet_by_index(idx):
        return sheet_list[idx]

    mock_book.sheet_by_index = sheet_by_index
    mock_book.sheet_names = MagicMock(return_value=list(sheets.keys()))

    return mock_book


def _mock_xlrd_sheet(
    name: str,
    rows: list[list[Any]],
    merged: list[tuple[int, int, int, int]] | None = None,
) -> MagicMock:
    """Create a mock xlrd Sheet with cell data."""
    nrows = len(rows)
    ncols = max((len(r) for r in rows), default=0)

    mock_sheet = MagicMock()
    mock_sheet.name = name
    mock_sheet.nrows = nrows
    mock_sheet.ncols = ncols
    mock_sheet.merged_cells = merged or []

    def cell_value(r, c):
        if r < len(rows) and c < len(rows[r]):
            return rows[r][c]
        return ""

    def cell_type(r, c):
        val = cell_value(r, c)
        if val is None or val == "":
            return xlrd.XL_CELL_EMPTY
        if isinstance(val, (int, float)):
            return xlrd.XL_CELL_NUMBER
        if isinstance(val, bool):
            return xlrd.XL_CELL_BOOLEAN
        return xlrd.XL_CELL_TEXT

    mock_sheet.cell_value = cell_value
    mock_sheet.cell_type = cell_type

    # Disable border detection by default
    mock_sheet.cell_xf_index = MagicMock(side_effect=Exception("no formatting"))

    return mock_sheet


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Constants
# ═══════════════════════════════════════════════════════════════════════════════

class TestConstants:
    def test_ole2_magic(self):
        assert OLE2_MAGIC == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        assert len(OLE2_MAGIC) == 8

    def test_zip_magic(self):
        assert ZIP_MAGIC == b"PK\x03\x04"

    def test_scan_limits(self):
        assert MAX_SCAN_ROWS == 1000
        assert MAX_SCAN_COLS == 100


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Converter
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsConverter:
    def test_format_name(self):
        c = XlsConverter()
        assert c.get_format_name() == "xls"

    def test_validate_empty(self):
        c = XlsConverter()
        assert c.validate(_make_ctx(b"")) is False

    def test_validate_too_short(self):
        c = XlsConverter()
        assert c.validate(_make_ctx(b"\xd0\xcf\x11")) is False

    def test_validate_wrong_magic(self):
        c = XlsConverter()
        assert c.validate(_make_ctx(b"PK\x03\x04" + b"\x00" * 100)) is False

    def test_validate_ole_magic(self):
        c = XlsConverter()
        data = OLE2_MAGIC + b"\x00" * 512
        assert c.validate(_make_ctx(data)) is True

    def test_convert_empty_raises(self):
        c = XlsConverter()
        with pytest.raises(ConversionError, match="Empty"):
            c.convert(_make_ctx(b""))

    def test_convert_bad_magic_raises(self):
        c = XlsConverter()
        with pytest.raises(ConversionError, match="magic"):
            c.convert(_make_ctx(b"not ole not ole not ole"))

    @patch("contextifier_new.handlers.xls.converter.xlrd")
    def test_convert_success(self, mock_xlrd):
        mock_book = MagicMock()
        mock_xlrd.open_workbook.return_value = mock_book

        c = XlsConverter()
        data = OLE2_MAGIC + b"\x00" * 100
        result = c.convert(_make_ctx(data))

        assert isinstance(result, XlsConvertedData)
        assert result.book is mock_book
        assert result.file_data == data
        mock_xlrd.open_workbook.assert_called_once()

    @patch("contextifier_new.handlers.xls.converter.xlrd")
    def test_convert_xlrd_error(self, mock_xlrd):
        mock_xlrd.open_workbook.side_effect = Exception("bad BIFF")

        c = XlsConverter()
        data = OLE2_MAGIC + b"\x00" * 100
        with pytest.raises(ConversionError, match="Failed"):
            c.convert(_make_ctx(data))

    def test_close_converted_data(self):
        book = MagicMock()
        cd = XlsConvertedData(book=book, file_data=b"")
        c = XlsConverter()
        c.close(cd)
        book.release_resources.assert_called_once()

    def test_close_none(self):
        c = XlsConverter()
        c.close(None)  # Should not raise

    def test_close_bare_book(self):
        book = MagicMock()
        book.release_resources = MagicMock()
        c = XlsConverter()
        c.close(book)
        book.release_resources.assert_called_once()

    def test_namedtuple_fields(self):
        cd = XlsConvertedData(book="B", file_data=b"D")
        assert cd[0] == "B"
        assert cd[1] == b"D"


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Preprocessor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsPreprocessor:
    def test_format_name(self):
        p = XlsPreprocessor()
        assert p.get_format_name() == "xls"

    def test_none_raises(self):
        p = XlsPreprocessor()
        with pytest.raises(PreprocessingError, match="None"):
            p.preprocess(None)

    def test_basic(self):
        book = _mock_xlrd_book({"Sales": [["A"]], "Info": [["B"]]})
        cd = XlsConvertedData(book=book, file_data=b"\x00")

        p = XlsPreprocessor()
        result = p.preprocess(cd)

        assert isinstance(result, PreprocessedData)
        assert result.content is book
        assert result.properties["sheet_count"] == 2
        assert "Sales" in result.properties["sheet_names"]
        assert result.resources["file_data"] == b"\x00"

    def test_bare_book(self):
        book = _mock_xlrd_book()
        p = XlsPreprocessor()
        result = p.preprocess(book)
        assert result.content is book
        assert result.resources["file_data"] == b""


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Metadata Extractor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsMetadataExtractor:
    def test_format_name(self):
        m = XlsMetadataExtractor()
        assert m.get_format_name() == "xls"

    def test_extract_none(self):
        m = XlsMetadataExtractor()
        result = m.extract(None)
        assert isinstance(result, DocumentMetadata)

    def test_extract_xlrd_user_name(self):
        book = _mock_xlrd_book(user_name="Author Name")
        m = XlsMetadataExtractor()
        result = m.extract(book)
        assert result.author == "Author Name"

    def test_extract_page_count(self):
        book = _mock_xlrd_book({"S1": [[]], "S2": [[]], "S3": [[]]})
        m = XlsMetadataExtractor()
        result = m.extract(book)
        assert result.page_count == 3

    @patch("contextifier_new.handlers.xls.metadata_extractor.olefile")
    def test_extract_from_ole(self, mock_olefile):
        mock_ole = MagicMock()
        mock_meta = MagicMock()
        mock_meta.title = "Test Title"
        mock_meta.subject = "Test Subject"
        mock_meta.author = "OLE Author"
        mock_meta.keywords = "kw1, kw2"
        mock_meta.comments = "Some comments"
        mock_meta.last_saved_by = "Modifier"
        mock_meta.create_time = datetime(2024, 1, 1)
        mock_meta.last_saved_time = datetime(2024, 6, 15)
        mock_meta.category = "Cat"
        mock_meta.revision_number = "5"
        mock_ole.get_metadata.return_value = mock_meta
        mock_olefile.OleFileIO.return_value = mock_ole

        book = _mock_xlrd_book()
        cd = XlsConvertedData(book=book, file_data=OLE2_MAGIC + b"\x00" * 100)

        m = XlsMetadataExtractor()
        result = m.extract(cd)

        assert result.title == "Test Title"
        assert result.author == "OLE Author"
        assert result.keywords == "kw1, kw2"
        assert result.create_time == datetime(2024, 1, 1)

    def test_extract_from_preprocessed_data(self):
        """Metadata extractor should unwrap PreprocessedData."""
        book = _mock_xlrd_book(user_name="PPD Author")
        ppd = PreprocessedData(
            content=book,
            raw_content=book,
            encoding="biff",
            resources={"file_data": b""},
            properties={},
        )
        m = XlsMetadataExtractor()
        result = m.extract(ppd)
        assert result.author == "PPD Author"

    def test_safe_str_bytes(self):
        m = XlsMetadataExtractor()
        assert m._safe_str(b"hello") == "hello"

    def test_safe_str_none(self):
        assert XlsMetadataExtractor._safe_str(None) is None

    def test_safe_str_empty(self):
        assert XlsMetadataExtractor._safe_str("") is None

    def test_safe_datetime_none(self):
        assert XlsMetadataExtractor._safe_datetime(None) is None

    def test_safe_datetime_value(self):
        dt = datetime(2024, 3, 15)
        assert XlsMetadataExtractor._safe_datetime(dt) == dt

    def test_safe_datetime_non_datetime(self):
        assert XlsMetadataExtractor._safe_datetime("not a date") is None


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Layout Detection
# ═══════════════════════════════════════════════════════════════════════════════

class TestLayoutRange:
    def test_properties(self):
        lr = LayoutRange(min_row=2, max_row=5, min_col=1, max_col=4)
        assert lr.rows == 4
        assert lr.cols == 4
        assert lr.cell_count == 16

    def test_contains(self):
        lr = LayoutRange(min_row=1, max_row=3, min_col=1, max_col=3)
        assert lr.contains(2, 2) is True
        assert lr.contains(4, 4) is False

    def test_overlaps(self):
        a = LayoutRange(1, 5, 1, 5)
        b = LayoutRange(3, 7, 3, 7)
        assert a.overlaps(b) is True

    def test_no_overlap(self):
        a = LayoutRange(1, 2, 1, 2)
        b = LayoutRange(4, 5, 4, 5)
        assert a.overlaps(b) is False

    def test_is_adjacent(self):
        a = LayoutRange(1, 3, 1, 3)
        b = LayoutRange(4, 6, 1, 3)
        assert a.is_adjacent(b) is True

    def test_merge_with(self):
        a = LayoutRange(1, 3, 1, 3)
        b = LayoutRange(5, 7, 2, 4)
        m = a.merge_with(b)
        assert m.min_row == 1 and m.max_row == 7
        assert m.min_col == 1 and m.max_col == 4


class TestLayoutDetection:
    def test_detect_empty(self):
        sheet = _mock_xlrd_sheet("E", [])
        sheet.nrows = 0
        sheet.ncols = 0
        assert layout_detect_range(sheet) is None

    def test_detect_single_cell(self):
        sheet = _mock_xlrd_sheet("S", [["Hello"]])
        lr = layout_detect_range(sheet)
        assert lr is not None
        assert lr.min_row == 1 and lr.max_row == 1
        assert lr.min_col == 1 and lr.max_col == 1

    def test_detect_multi(self):
        sheet = _mock_xlrd_sheet("M", [
            ["A", "B"],
            ["C", "D"],
            ["E", "F"],
        ])
        lr = layout_detect_range(sheet)
        assert lr is not None
        assert lr.rows == 3
        assert lr.cols == 2

    def test_detect_offset(self):
        """Data starting at row 1, col 1 (0-based B2)."""
        rows = [["", ""], ["", "X", "Y"], ["", "Z", "W"]]
        sheet = _mock_xlrd_sheet("O", rows)
        lr = layout_detect_range(sheet)
        assert lr is not None
        assert lr.min_row == 2 and lr.min_col == 2

    def test_object_detect_basic(self):
        sheet = _mock_xlrd_sheet("T", [["A", "B"], ["C", "D"]])
        regions = object_detect(sheet, None)
        assert len(regions) >= 1

    def test_object_detect_empty(self):
        sheet = _mock_xlrd_sheet("E", [])
        sheet.nrows = 0
        sheet.ncols = 0
        assert object_detect(sheet, None) == []


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Table Conversion
# ═══════════════════════════════════════════════════════════════════════════════

class TestTableConversion:
    def test_markdown_basic(self):
        sheet = _mock_xlrd_sheet("S", [["Name", "Age"], ["Alice", 30]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 2)

        md = convert_region_to_markdown(sheet, book, region)
        assert "Name" in md
        assert "Age" in md
        assert "Alice" in md
        assert "30" in md
        assert "---" in md
        assert "|" in md

    def test_markdown_pipe_escaping(self):
        sheet = _mock_xlrd_sheet("S", [["A|B"]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 1, 1, 1)

        md = convert_region_to_markdown(sheet, book, region)
        assert "\\|" in md

    def test_html_basic(self):
        sheet = _mock_xlrd_sheet("H", [["Header"], ["Data"]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 1)

        html = convert_region_to_html(sheet, book, region)
        assert "<table>" in html
        assert "<th>" in html
        assert "Header" in html
        assert "<td>" in html
        assert "Data" in html

    def test_html_escaping(self):
        sheet = _mock_xlrd_sheet("H", [["<script>"]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 1, 1, 1)

        html = convert_region_to_html(sheet, book, region)
        assert "&lt;script&gt;" in html
        assert "<script>" not in html.replace("&lt;script&gt;", "")

    def test_html_with_merged(self):
        """Merged cells should produce rowspan/colspan."""
        sheet = _mock_xlrd_sheet(
            "M",
            [["Merged Header", ""], ["V1", "V2"]],
            merged=[(0, 1, 0, 2)],  # row 0, cols 0-1 merged (half-open)
        )
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 2)

        html = convert_region_to_html(sheet, book, region)
        assert "colspan='2'" in html

    def test_auto_select_md_no_merge(self):
        sheet = _mock_xlrd_sheet("S", [["X"], ["Y"]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 1)

        text = convert_sheet_to_text(sheet, book, region)
        assert "<table>" not in text
        assert "|" in text

    def test_auto_select_html_with_merge(self):
        sheet = _mock_xlrd_sheet(
            "S", [["A", ""], ["B", "C"]],
            merged=[(0, 1, 0, 2)],
        )
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 2)

        text = convert_sheet_to_text(sheet, book, region)
        assert "<table>" in text

    def test_convert_region_to_table(self):
        sheet = _mock_xlrd_sheet("T", [["H1", "H2"], ["V1", "V2"]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 2)

        td = convert_region_to_table(sheet, book, region)
        assert td is not None
        assert isinstance(td, TableData)
        assert td.num_rows == 2
        assert td.num_cols == 2

    def test_convert_region_to_table_empty(self):
        sheet = _mock_xlrd_sheet("E", [["", ""], ["", ""]])
        book = _mock_xlrd_book()
        region = LayoutRange(1, 2, 1, 2)

        td = convert_region_to_table(sheet, book, region)
        assert td is None


class TestFormatCell:
    def test_text(self):
        sheet = _mock_xlrd_sheet("S", [["Hello"]])
        book = _mock_xlrd_book()
        assert _format_cell(sheet, book, 0, 0) == "Hello"

    def test_number_int(self):
        sheet = _mock_xlrd_sheet("S", [[42.0]])
        book = _mock_xlrd_book()
        assert _format_cell(sheet, book, 0, 0) == "42"

    def test_number_float(self):
        sheet = _mock_xlrd_sheet("S", [[3.14]])
        book = _mock_xlrd_book()
        assert _format_cell(sheet, book, 0, 0) == "3.14"

    def test_empty(self):
        sheet = _mock_xlrd_sheet("S", [[""]])
        book = _mock_xlrd_book()
        assert _format_cell(sheet, book, 0, 0) == ""

    def test_boolean_true(self):
        sheet = _mock_xlrd_sheet("S", [[True]])
        # Override cell_type to return BOOLEAN
        sheet.cell_type = MagicMock(return_value=xlrd.XL_CELL_BOOLEAN)
        sheet.cell_value = MagicMock(return_value=True)
        book = _mock_xlrd_book()
        assert _format_cell(sheet, book, 0, 0) == "TRUE"


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Content Extractor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsContentExtractor:
    def test_format_name(self):
        ext = XlsContentExtractor()
        assert ext.get_format_name() == "xls"

    def test_extract_text_basic(self):
        book = _mock_xlrd_book({"Sheet1": [["Hello", "World"], ["A", "B"]]})
        ppd = PreprocessedData(
            content=book,
            raw_content=book,
            encoding="biff",
            resources={"file_data": b""},
            properties={"sheet_names": ["Sheet1"]},
        )
        ext = XlsContentExtractor()
        text = ext.extract_text(ppd)

        assert "[Sheet: Sheet1]" in text
        assert "Hello" in text
        assert "World" in text

    def test_extract_text_multi_sheet(self):
        book = _mock_xlrd_book({
            "Sales": [["Product", "Revenue"], ["A", 100]],
            "Info": [["Key", "Value"], ["Name", "Test"]],
        })
        ppd = PreprocessedData(
            content=book,
            raw_content=book,
            encoding="biff",
            resources={"file_data": b""},
            properties={"sheet_names": ["Sales", "Info"]},
        )
        ext = XlsContentExtractor()
        text = ext.extract_text(ppd)

        assert "[Sheet: Sales]" in text
        assert "[Sheet: Info]" in text
        assert "Product" in text
        assert "Name" in text

    def test_extract_text_none(self):
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="biff",
            resources={},
            properties={},
        )
        ext = XlsContentExtractor()
        assert ext.extract_text(ppd) == ""

    def test_extract_text_with_tag_service(self):
        book = _mock_xlrd_book({"Sheet1": [["X"]]})
        tag_svc = MagicMock()
        tag_svc.make_sheet_tag.return_value = "[CUSTOM: Sheet1]"

        ppd = PreprocessedData(
            content=book, raw_content=book, encoding="biff",
            resources={}, properties={},
        )
        ext = XlsContentExtractor(tag_service=tag_svc)
        text = ext.extract_text(ppd)
        assert "[CUSTOM: Sheet1]" in text

    def test_extract_tables(self):
        book = _mock_xlrd_book({"S": [["H1", "H2"], ["V1", "V2"]]})
        ppd = PreprocessedData(
            content=book, raw_content=book, encoding="biff",
            resources={}, properties={},
        )
        ext = XlsContentExtractor()
        tables = ext.extract_tables(ppd)
        assert len(tables) >= 1

    def test_extract_images_always_empty(self):
        ext = XlsContentExtractor()
        ppd = PreprocessedData(
            content=None, raw_content=None, encoding="biff",
            resources={}, properties={},
        )
        assert ext.extract_images(ppd) == []

    def test_extract_charts_always_empty(self):
        ext = XlsContentExtractor()
        ppd = PreprocessedData(
            content=None, raw_content=None, encoding="biff",
            resources={}, properties={},
        )
        assert ext.extract_charts(ppd) == []


# ═══════════════════════════════════════════════════════════════════════════════
# 8. Handler
# ═══════════════════════════════════════════════════════════════════════════════

class TestXLSHandler:
    def test_supported_extensions(self):
        h = XLSHandler(ProcessingConfig())
        assert h.supported_extensions == frozenset({"xls"})

    def test_handler_name(self):
        h = XLSHandler(ProcessingConfig())
        assert h.handler_name == "XLS Handler"

    def test_creates_converter(self):
        h = XLSHandler(ProcessingConfig())
        assert isinstance(h.create_converter(), XlsConverter)

    def test_creates_preprocessor(self):
        h = XLSHandler(ProcessingConfig())
        assert isinstance(h.create_preprocessor(), XlsPreprocessor)

    def test_creates_metadata_extractor(self):
        h = XLSHandler(ProcessingConfig())
        assert isinstance(h.create_metadata_extractor(), XlsMetadataExtractor)

    def test_creates_content_extractor(self):
        h = XLSHandler(ProcessingConfig())
        assert isinstance(h.create_content_extractor(), XlsContentExtractor)


# ═══════════════════════════════════════════════════════════════════════════════
# 9. Delegation
# ═══════════════════════════════════════════════════════════════════════════════

class TestDelegation:
    def test_zip_delegates_to_xlsx(self):
        """If a .xls file has ZIP magic, delegate to xlsx handler."""
        import openpyxl as oxl
        wb = oxl.Workbook()
        wb.active["A1"] = "Delegated"
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        data = buf.getvalue()
        handler = XLSHandler(ProcessingConfig())

        # Set up mock registry so delegation works
        mock_registry = MagicMock()
        mock_xlsx_handler = MagicMock()
        mock_xlsx_handler.process.return_value = ExtractionResult(text="Delegated content")
        mock_registry.get_handler.return_value = mock_xlsx_handler
        handler._handler_registry = mock_registry

        ctx = _make_ctx(data, ext="xls", name="actually_xlsx.xls")
        result = handler._check_delegation(ctx)

        assert result is not None
        assert result.text == "Delegated content"
        mock_registry.get_handler.assert_called_with("xlsx")

    def test_ole_no_delegation(self):
        """OLE2 magic → _check_delegation returns None (process as XLS)."""
        handler = XLSHandler(ProcessingConfig())
        data = OLE2_MAGIC + b"\x00" * 100
        ctx = _make_ctx(data)
        result = handler._check_delegation(ctx)
        assert result is None

    def test_empty_no_delegation(self):
        handler = XLSHandler(ProcessingConfig())
        ctx = _make_ctx(b"")
        result = handler._check_delegation(ctx)
        assert result is None


# ═══════════════════════════════════════════════════════════════════════════════
# 10. Full Pipeline & Edge Cases
# ═══════════════════════════════════════════════════════════════════════════════

class TestFullPipeline:
    @patch("contextifier_new.handlers.xls.converter.xlrd")
    def test_pipeline_basic(self, mock_xlrd_mod):
        """Test the full pipeline with a mocked xlrd."""
        book = _mock_xlrd_book({"Sheet1": [["Header1", "Header2"], ["Val1", "Val2"]]})
        mock_xlrd_mod.open_workbook.return_value = book

        handler = XLSHandler(ProcessingConfig())
        data = OLE2_MAGIC + b"\x00" * 100
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert isinstance(result, ExtractionResult)
        assert "Header1" in result.text
        assert "Val1" in result.text

    @patch("contextifier_new.handlers.xls.converter.xlrd")
    def test_pipeline_multi_sheet(self, mock_xlrd_mod):
        book = _mock_xlrd_book({
            "Sales": [["P", "Rev"], ["A", 100]],
            "Meta": [["K", "V"], ["N", "T"]],
        })
        mock_xlrd_mod.open_workbook.return_value = book

        handler = XLSHandler(ProcessingConfig())
        data = OLE2_MAGIC + b"\x00" * 100
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert "[Sheet: Sales]" in result.text
        assert "[Sheet: Meta]" in result.text

    @patch("contextifier_new.handlers.xls.converter.xlrd")
    def test_pipeline_metadata(self, mock_xlrd_mod):
        book = _mock_xlrd_book(user_name="Pipeline Author")
        mock_xlrd_mod.open_workbook.return_value = book

        handler = XLSHandler(ProcessingConfig())
        data = OLE2_MAGIC + b"\x00" * 100
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert result.metadata is not None
        assert result.metadata.author == "Pipeline Author"


class TestEdgeCases:
    def test_empty_sheet(self):
        book = _mock_xlrd_book({"Empty": []})
        ws = book.sheet_by_index(0)
        ws.nrows = 0
        ws.ncols = 0

        ppd = PreprocessedData(
            content=book, raw_content=book, encoding="biff",
            resources={}, properties={},
        )
        ext = XlsContentExtractor()
        text = ext.extract_text(ppd)
        assert "[Sheet: Empty]" in text

    def test_single_cell(self):
        book = _mock_xlrd_book({"One": [["Solo"]]})
        ppd = PreprocessedData(
            content=book, raw_content=book, encoding="biff",
            resources={}, properties={},
        )
        ext = XlsContentExtractor()
        text = ext.extract_text(ppd)
        assert "Solo" in text

    def test_layout_range_single(self):
        lr = LayoutRange(1, 1, 1, 1)
        assert lr.rows == 1
        assert lr.cols == 1
        assert lr.cell_count == 1

    def test_converter_close_error_ignored(self):
        book = MagicMock()
        book.release_resources.side_effect = Exception("oops")
        cd = XlsConvertedData(book=book, file_data=b"")
        c = XlsConverter()
        c.close(cd)  # Should not raise

    def test_content_extractor_bad_preprocessed(self):
        ext = XlsContentExtractor()
        # Pass something without nsheets
        ppd = PreprocessedData(
            content="not a book", raw_content="not a book", encoding="biff",
            resources={}, properties={},
        )
        result = ext.extract_text(ppd)
        assert result == ""

    def test_make_sheet_tag_fallback(self):
        ext = XlsContentExtractor(tag_service=None)
        assert ext._make_sheet_tag("MySheet") == "[Sheet: MySheet]"

    def test_make_sheet_tag_service_error(self):
        svc = MagicMock()
        svc.make_sheet_tag.side_effect = Exception("fail")
        ext = XlsContentExtractor(tag_service=svc)
        assert ext._make_sheet_tag("S") == "[Sheet: S]"
