# tests/unit/raw/test_xlsx_raw.py
"""XLSX raw model — surgical cell edits under the byte-preservation contract.

The headline: an edit through :class:`XlsxRawDocument` re-serializes
exactly the worksheet it touched (plus ``xl/workbook.xml`` when a
``calcPr`` flag is needed) — everything an openpyxl round-trip destroys
(chart style parts, custom XML, sparkline extLst blocks) survives
byte-identical.

Fixtures are built with openpyxl and then, where openpyxl cannot express
a feature (shared strings, formula cached values, foreign parts), the
package is patched at the ZIP level to look like an Excel-saved file.
Every patch asserts its anchor string so openpyxl output drift fails
loudly instead of silently weakening a test.
"""

from __future__ import annotations

import io
import re
import warnings
import zipfile

import openpyxl
import pytest
from lxml import etree
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName

from contextifier.raw import open_raw
from contextifier.raw.xlsx import (
    XlsxRawDocument,
    col_index_to_letters,
    col_letters_to_index,
    parse_ref,
)
from contextifier.raw.xmlpart import NS

SHEET1 = "xl/worksheets/sheet1.xml"
SHEET2 = "xl/worksheets/sheet2.xml"
WORKBOOK = "xl/workbook.xml"
WORKBOOK_RELS = "xl/_rels/workbook.xml.rels"
CONTENT_TYPES = "[Content_Types].xml"
SHARED_STRINGS = "xl/sharedStrings.xml"


# -- fixture builders --------------------------------------------------------


def _build_workbook(with_image: bool = False) -> bytes:
    """Two sheets, strings/numbers/bools/formula, merge, number format,
    a bar chart, a defined name — and optionally an image."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["C1"] = True
    ws["D1"] = 1234.5
    ws["D1"].number_format = "#,##0.00"
    ws["A2"] = "East"
    ws["B2"] = 100
    ws["A3"] = "West"
    ws["B3"] = 250.5
    ws["B4"] = "=SUM(B2:B3)"
    ws["A5"] = "Merged"
    ws.merge_cells("A5:B5")
    chart = BarChart()
    chart.add_data(
        Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(chart, "E7")
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="Sales!$A$1")
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "hello"
    if with_image:
        from openpyxl.drawing.image import Image as XlImage
        from PIL import Image as PilImage

        ibuf = io.BytesIO()
        PilImage.new("RGB", (4, 4), (200, 30, 30)).save(ibuf, format="PNG")
        ibuf.seek(0)
        ws.add_image(XlImage(ibuf), "F10")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parts(data: bytes) -> dict[str, bytes]:
    z = zipfile.ZipFile(io.BytesIO(data))
    return {n: z.read(n) for n in z.namelist()}


def _rezip(
    data: bytes,
    *,
    replace: dict[str, bytes] | None = None,
    add: dict[str, bytes] | None = None,
) -> bytes:
    z = zipfile.ZipFile(io.BytesIO(data))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as znew:
        for n in z.namelist():
            znew.writestr(n, (replace or {}).get(n, z.read(n)))
        for n, content in (add or {}).items():
            znew.writestr(n, content)
    return out.getvalue()


def _diff_parts(a: bytes, b: bytes) -> set[str]:
    pa, pb = _parts(a), _parts(b)
    assert set(pa) == set(pb), "part list drift"
    return {n for n in pa if pa[n] != pb[n]}


def _load_wb(data: bytes) -> openpyxl.Workbook:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # openpyxl warns on features it drops
        return openpyxl.load_workbook(io.BytesIO(data))


def _sheet_root(data: bytes, part: str):
    return etree.fromstring(_parts(data)[part])


def _cell_el(data: bytes, part: str, ref: str):
    for c in _sheet_root(data, part).findall("s:sheetData/s:row/s:c", NS):
        if c.get("r") == ref:
            return c
    return None


def _excel_flavoured(base: bytes) -> bytes:
    """Patch openpyxl output to look Excel-saved: a shared-string cell,
    a cached formula value, and a calcPr *without* fullCalcOnLoad."""
    parts = _parts(base)
    s1 = parts[SHEET1]
    anchor = b"<f>SUM(B2:B3)</f><v></v>"
    assert anchor in s1, "openpyxl formula serialization drifted"
    s1 = s1.replace(anchor, b"<f>SUM(B2:B3)</f><v>350.5</v>")
    anchor = b'<c r="A1" t="inlineStr"><is><t>Region</t></is></c>'
    assert anchor in s1, "openpyxl inline-string serialization drifted"
    s1 = s1.replace(anchor, b'<c r="A1" t="s"><v>0</v></c>')
    wb = re.sub(rb"<calcPr [^>]*/>", b'<calcPr calcId="1"/>', parts[WORKBOOK])
    assert b"fullCalcOnLoad" not in wb
    ct = parts[CONTENT_TYPES].replace(
        b"</Types>",
        b'<Override PartName="/xl/sharedStrings.xml" ContentType="application/'
        b'vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        b"</Types>",
    )
    rels = parts[WORKBOOK_RELS].replace(
        b"</Relationships>",
        b'<Relationship Id="rId97" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/sharedStrings" '
        b'Target="sharedStrings.xml"/></Relationships>',
    )
    sst = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        b' count="1" uniqueCount="1">'
        b'<si><r><t xml:space="preserve">Reg</t></r><r><t>ion</t></r></si></sst>'
    )
    return _rezip(
        base,
        replace={
            SHEET1: s1,
            WORKBOOK: wb,
            CONTENT_TYPES: ct,
            WORKBOOK_RELS: rels,
        },
        add={SHARED_STRINGS: sst},
    )


_SPARKLINE_EXT = (
    b'<extLst><ext xmlns:x14="http://schemas.microsoft.com/office/'
    b'spreadsheetml/2009/9/main" uri="{05C60535-1F16-4fd2-B633-F4F36F0B64E0}">'
    b'<x14:sparklineGroups xmlns:xm="http://schemas.microsoft.com/office/'
    b'excel/2006/main"><x14:sparklineGroup displayEmptyCellsAs="gap">'
    b'<x14:colorSeries rgb="FF376092"/><x14:sparklines><x14:sparkline>'
    b"<xm:f>Sales!B2:B3</xm:f><xm:sqref>B2</xm:sqref></x14:sparkline>"
    b"</x14:sparklines></x14:sparklineGroup></x14:sparklineGroups>"
    b"</ext></extLst>"
)

FOREIGN_PARTS = [
    "xl/charts/style1.xml",
    "xl/charts/colors1.xml",
    "xl/charts/_rels/chart1.xml.rels",
    "customXml/item1.xml",
    "customXml/itemProps1.xml",
    "customXml/_rels/item1.xml.rels",
]


def _inject_foreign_features(base: bytes) -> bytes:
    """Add the parts openpyxl round-trips destroy: chart style/colors,
    custom XML (with rels + content types), and a sparkline extLst
    inside the Notes worksheet."""
    parts = _parts(base)
    ct = parts[CONTENT_TYPES].replace(
        b"</Types>",
        b'<Override PartName="/xl/charts/style1.xml" ContentType="application/'
        b'vnd.ms-office.chartstyle+xml"/>'
        b'<Override PartName="/xl/charts/colors1.xml" ContentType="application/'
        b'vnd.ms-office.chartcolorstyle+xml"/>'
        b'<Override PartName="/customXml/itemProps1.xml" ContentType="application/'
        b'vnd.openxmlformats-officedocument.customXmlProperties+xml"/>'
        b"</Types>",
    )
    wb_rels = parts[WORKBOOK_RELS].replace(
        b"</Relationships>",
        b'<Relationship Id="rId98" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/customXml" '
        b'Target="../customXml/item1.xml"/></Relationships>',
    )
    s2 = parts[SHEET2]
    assert b"</worksheet>" in s2
    s2 = s2.replace(b"</worksheet>", _SPARKLINE_EXT + b"</worksheet>")
    chart_rels = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        b'relationships">'
        b'<Relationship Id="rId1" Type="http://schemas.microsoft.com/office/'
        b'2011/relationships/chartStyle" Target="style1.xml"/>'
        b'<Relationship Id="rId2" Type="http://schemas.microsoft.com/office/'
        b'2011/relationships/chartColorStyle" Target="colors1.xml"/>'
        b"</Relationships>"
    )
    item_rels = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        b'relationships">'
        b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/customXmlProps" '
        b'Target="itemProps1.xml"/></Relationships>'
    )
    return _rezip(
        base,
        replace={CONTENT_TYPES: ct, WORKBOOK_RELS: wb_rels, SHEET2: s2},
        add={
            "xl/charts/style1.xml": (
                b'<cs:chartStyle xmlns:cs="http://schemas.microsoft.com/office/'
                b'drawing/2012/chartStyle" id="201"/>'
            ),
            "xl/charts/colors1.xml": (
                b'<cs:colorStyle xmlns:cs="http://schemas.microsoft.com/office/'
                b'drawing/2012/chartStyle" meth="cycle" id="10"/>'
            ),
            "customXml/item1.xml": (
                '<meta xmlns="urn:example:meta"><owner>홍길동</owner></meta>'
            ).encode("utf-8"),
            "customXml/itemProps1.xml": (
                b'<ds:datastoreItem xmlns:ds="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/customXml" ds:itemID="{A1B2C3D4-0000-0000-'
                b'0000-000000000001}"/>'
            ),
            "customXml/_rels/item1.xml.rels": item_rels,
            "xl/charts/_rels/chart1.xml.rels": chart_rels,
        },
    )


# -- reading ------------------------------------------------------------------


class TestReadModel:
    def test_open_raw_dispatch(self):
        raw = open_raw(_build_workbook())  # bytes: format is sniffed
        assert isinstance(raw, XlsxRawDocument)
        assert raw.format == "xlsx"

    def test_sheet_accessors(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        assert raw.sheet_names == ["Sales", "Notes"]
        assert raw.sheets[0].name == "Sales"
        assert raw.sheets["Notes"].part_name == SHEET2
        assert len(raw.sheets) == 2
        assert [s.name for s in raw.sheets] == ["Sales", "Notes"]
        assert "Sales" in raw.sheets and "Nope" not in raw.sheets
        with pytest.raises(KeyError, match="Nope"):
            raw.sheets["Nope"]
        with pytest.raises(IndexError):
            raw.sheets[9]

    def test_values_and_types(self):
        raw = open_raw(_excel_flavoured(_build_workbook()), extension="xlsx")
        sales = raw.sheets["Sales"]
        assert sales.get_cell("A1") == "Region"  # shared string, 2 runs
        v = sales.get_cell("B2")
        assert v == 100 and isinstance(v, int)
        v = sales.get_cell("B3")
        assert v == 250.5 and isinstance(v, float)
        assert sales.get_cell("C1") is True
        assert sales.get_cell("A5") == "Merged"  # inline string
        assert raw.sheets["Notes"].get_cell("A1") == "hello"

    def test_formula_cached_value_and_text(self):
        raw = open_raw(_excel_flavoured(_build_workbook()), extension="xlsx")
        sales = raw.sheets["Sales"]
        assert sales.get_cell("B4") == 350.5  # the cached result
        assert sales.get_formula("B4") == "SUM(B2:B3)"
        assert sales.get_formula("B3") is None
        assert sales.get_formula("Z99") is None

    def test_dimensions_and_merged_ranges(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        assert raw.sheets["Sales"].dimensions == (5, 4)
        assert raw.sheets["Sales"].merged_ranges == ["A5:B5"]
        assert raw.sheets["Notes"].dimensions == (1, 1)
        assert raw.sheets["Notes"].merged_ranges == []

    def test_defined_names(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        assert raw.defined_names == {"MyRange": "Sales!$A$1"}

    def test_chart_part_names(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        assert raw.chart_part_names == ["xl/charts/chart1.xml"]

    def test_iter_rows(self):
        raw = open_raw(_excel_flavoured(_build_workbook()), extension="xlsx")
        cells = dict(raw.sheets["Sales"].iter_rows())
        assert cells["A1"] == "Region"
        assert cells["B3"] == 250.5
        window = dict(raw.sheets["Sales"].iter_rows(min_row=2, max_row=3))
        assert set(window) == {"A2", "B2", "A3", "B3"}

    def test_missing_cells_are_none(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        sales = raw.sheets["Sales"]
        assert sales.get_cell("Z99") is None  # row doesn't exist
        assert sales.get_cell("D5") is None  # row exists, cell doesn't
        with pytest.raises(ValueError, match="reference"):
            sales.get_cell("not-a-ref")

    def test_shared_string_ref_without_table_is_none(self):
        """No sharedStrings part at all: constructor stays resilient and
        a dangling t="s" cell reads as None instead of raising."""
        parts = _parts(_build_workbook())
        s1 = parts[SHEET1]
        anchor = b'<c r="A1" t="inlineStr"><is><t>Region</t></is></c>'
        assert anchor in s1
        data = _rezip(
            _build_workbook(),
            replace={SHEET1: s1.replace(anchor, b'<c r="A1" t="s"><v>0</v></c>')},
        )
        raw = open_raw(data, extension="xlsx")
        assert raw.sheets["Sales"].get_cell("A1") is None

    def test_ref_helpers(self):
        assert col_letters_to_index("A") == 1
        assert col_letters_to_index("AA") == 27
        assert col_index_to_letters(28) == "AB"
        for i in (1, 26, 27, 52, 703):
            assert col_letters_to_index(col_index_to_letters(i)) == i
        assert parse_ref("$B$3") == (3, 2)
        with pytest.raises(ValueError):
            parse_ref("B0")


# -- surgical writes -----------------------------------------------------------


class TestSurgicalWrites:
    def test_single_set_cell_diffs_exactly_one_part(self):
        base = _build_workbook()
        # openpyxl already emits fullCalcOnLoad, so workbook.xml stays clean
        assert b'fullCalcOnLoad="1"' in _parts(base)[WORKBOOK]
        raw = open_raw(base, extension="xlsx")
        raw.sheets["Sales"].set_cell("B3", 999)
        out = raw.to_bytes()

        assert _diff_parts(base, out) == {SHEET1}
        assert _parts(out)[CONTENT_TYPES] == _parts(base)[CONTENT_TYPES]

        wb = _load_wb(out)
        assert wb["Sales"]["B3"].value == 999
        assert wb["Sales"]["A1"].value == "Region"  # neighbors untouched
        assert wb["Sales"]["D1"].number_format == "#,##0.00"  # styles intact

    def test_style_attribute_preserved_on_overwrite(self):
        base = _build_workbook()
        orig_style = _cell_el(base, SHEET1, "D1").get("s")
        assert orig_style is not None  # the number format style index
        raw = open_raw(base, extension="xlsx")
        raw.sheets["Sales"].set_cell("D1", 777)
        out = raw.to_bytes()
        d1 = _cell_el(out, SHEET1, "D1")
        assert d1.get("s") == orig_style
        assert d1.get("t") is None  # numeric: no type attribute
        wb = _load_wb(out)
        assert wb["Sales"]["D1"].value == 777
        assert wb["Sales"]["D1"].number_format == "#,##0.00"

    def test_set_none_clears_value_keeps_cell_and_style(self):
        base = _build_workbook()
        orig_style = _cell_el(base, SHEET1, "D1").get("s")
        raw = open_raw(base, extension="xlsx")
        raw.sheets["Sales"].set_cell("D1", None)
        assert raw.sheets["Sales"].get_cell("D1") is None
        out = raw.to_bytes()
        d1 = _cell_el(out, SHEET1, "D1")
        assert d1 is not None and d1.get("s") == orig_style
        assert len(d1) == 0  # no v / is children left

    def test_bool_and_float_write_roundtrip(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        sales = raw.sheets["Sales"]
        sales.set_cell("C2", False)
        sales.set_cell("C3", 0.125)
        wb = _load_wb(raw.to_bytes())
        assert wb["Sales"]["C2"].value is False
        assert wb["Sales"]["C3"].value == 0.125

    def test_unsupported_type_raises(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        with pytest.raises(TypeError, match="dict"):
            raw.sheets["Sales"].set_cell("A1", {"no": "way"})


# -- the killer test: foreign-feature survival ---------------------------------


class TestForeignFeatureSurvival:
    def test_injected_parts_and_extlst_survive_raw_edit(self):
        injected = _inject_foreign_features(_build_workbook())
        raw = open_raw(injected, extension="xlsx")
        raw.sheets["Sales"].set_cell("B3", 777)
        out = raw.save()

        # the ONLY part that differs is the edited worksheet
        assert _diff_parts(injected, out) == {SHEET1}
        out_parts, in_parts = _parts(out), _parts(injected)
        for name in FOREIGN_PARTS + [SHEET2, CONTENT_TYPES, WORKBOOK_RELS]:
            assert out_parts[name] == in_parts[name], f"{name} was not preserved"
        assert b"sparklineGroups" in out_parts[SHEET2]
        assert "홍길동".encode("utf-8") in out_parts["customXml/item1.xml"]
        # and the edit itself landed
        assert _load_wb(out)["Sales"]["B3"].value == 777

    def test_openpyxl_roundtrip_destroys_the_same_features(self):
        """The contrast: the exact features we preserve, openpyxl drops."""
        injected = _inject_foreign_features(_build_workbook())
        wb = _load_wb(injected)
        buf = io.BytesIO()
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb.save(buf)
        survivors = set(_parts(buf.getvalue()))
        assert "xl/charts/style1.xml" not in survivors
        assert "xl/charts/colors1.xml" not in survivors
        assert "customXml/item1.xml" not in survivors
        s2 = _parts(buf.getvalue()).get(SHEET2, b"")
        assert b"sparklineGroups" not in s2

    def test_sparkline_extlst_survives_editing_its_own_sheet(self):
        """Even when the sparkline's OWN sheet is edited (so its XML is
        re-serialized), the extLst rides along in the tree."""
        injected = _inject_foreign_features(_build_workbook())
        raw = open_raw(injected, extension="xlsx")
        raw.sheets["Notes"].set_cell("A2", "edited")
        out = raw.to_bytes()
        s2 = _parts(out)[SHEET2]
        assert b"sparklineGroups" in s2
        assert b"Sales!B2:B3" in s2
        assert _load_wb(out)["Notes"]["A2"].value == "edited"

    def test_image_media_part_survives(self):
        pytest.importorskip("PIL")
        base = _build_workbook(with_image=True)
        assert "xl/media/image1.png" in _parts(base)
        raw = open_raw(base, extension="xlsx")
        raw.sheets["Sales"].set_cell("B3", 1)
        out = raw.to_bytes()
        assert _parts(out)["xl/media/image1.png"] == _parts(base)["xl/media/image1.png"]


# -- formulas -------------------------------------------------------------------


class TestFormulaOverwrite:
    def test_overwrite_removes_formula_and_forces_recalc(self):
        data = _excel_flavoured(_build_workbook())  # calcPr WITHOUT fullCalcOnLoad
        raw = open_raw(data, extension="xlsx")
        sales = raw.sheets["Sales"]
        assert sales.get_formula("B4") == "SUM(B2:B3)"
        sales.set_cell("B4", 42)
        assert sales.get_formula("B4") is None
        assert sales.get_cell("B4") == 42
        out = raw.to_bytes()
        assert b"<f>" not in _parts(out)[SHEET1]  # formula gone
        assert re.search(rb'<calcPr[^>]*fullCalcOnLoad="1"', _parts(out)[WORKBOOK])
        assert _load_wb(out)["Sales"]["B4"].value == 42  # the literal

    def test_calc_pr_created_after_sheets_when_absent(self):
        base = _excel_flavoured(_build_workbook())
        stripped = re.sub(rb"<calcPr[^>]*/>", b"", _parts(base)[WORKBOOK])
        assert b"calcPr" not in stripped
        data = _rezip(base, replace={WORKBOOK: stripped})
        raw = open_raw(data, extension="xlsx")
        raw.sheets["Sales"].set_cell("B4", 5)
        out = raw.to_bytes()
        out_wb = _parts(out)[WORKBOOK]
        assert b'<calcPr fullCalcOnLoad="1"/>' in out_wb
        assert out_wb.index(b"<calcPr") > out_wb.index(b"</definedNames>")
        _load_wb(out)  # still a valid workbook

    def test_plain_edit_with_formulas_elsewhere_forces_recalc(self):
        """B4's =SUM depends on B2 — editing B2 must trigger recalc-on-load
        even though B2 itself has no formula."""
        data = _excel_flavoured(_build_workbook())
        raw = open_raw(data, extension="xlsx")
        raw.sheets["Sales"].set_cell("B2", 999)
        out = raw.to_bytes()
        assert re.search(rb'fullCalcOnLoad="1"', _parts(out)[WORKBOOK])
        assert _diff_parts(data, out) == {SHEET1, WORKBOOK}


# -- appending ------------------------------------------------------------------


class TestAppendRows:
    def test_append_rows_values_and_dimension(self):
        base = _build_workbook()
        raw = open_raw(base, extension="xlsx")
        raw.sheets["Sales"].append_rows([["North", 300], ["South", 400.5, True]])
        out = raw.to_bytes()

        wb = _load_wb(out)
        ws = wb["Sales"]
        assert ws["A6"].value == "North" and ws["B6"].value == 300
        assert ws["A7"].value == "South" and ws["B7"].value == 400.5
        assert ws["C7"].value is True

        assert b'ref="A1:D7"' in _parts(out)[SHEET1]  # dimension extended
        raw2 = open_raw(out, extension="xlsx")
        assert raw2.sheets["Sales"].dimensions == (7, 4)
        assert raw2.sheets["Sales"].get_cell("B7") == 400.5

    def test_append_sparse_rows_skips_none(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        raw.sheets["Notes"].append_rows([[None, "only-b"]])
        out = raw.to_bytes()
        assert _cell_el(out, SHEET2, "A2") is None  # sparse: A2 not stored
        assert _load_wb(out)["Notes"]["B2"].value == "only-b"


# -- inline strings ---------------------------------------------------------------


class TestInlineStrings:
    def test_korean_and_xml_specials_roundtrip(self):
        base = _build_workbook()
        raw = open_raw(base, extension="xlsx")
        text = "한글 데이터 <&> \"quote\" 'tick' 100%"
        raw.sheets["Notes"].set_cell("A2", text)
        out = raw.to_bytes()
        assert _load_wb(out)["Notes"]["A2"].value == text
        raw2 = open_raw(out, extension="xlsx")
        assert raw2.sheets["Notes"].get_cell("A2") == text
        # inlineStr writes never create a shared-string table
        assert SHARED_STRINGS not in _parts(out)

    def test_leading_trailing_space_preserved(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        raw.sheets["Notes"].set_cell("A3", "  padded  ")
        out = raw.to_bytes()
        assert _load_wb(out)["Notes"]["A3"].value == "  padded  "


# -- ordering invariants ------------------------------------------------------------


class TestOrderingInvariants:
    def test_new_rows_inserted_sorted_by_r(self):
        """Excel requires ascending @r; creating a row between existing
        rows must keep the order."""
        raw = open_raw(_build_workbook(), extension="xlsx")
        notes = raw.sheets["Notes"]
        notes.set_cell("A5", "five")
        notes.set_cell("A3", "three")  # lands BETWEEN rows 1 and 5
        out = raw.to_bytes()
        rows = [
            r.get("r")
            for r in _sheet_root(out, SHEET2).findall("s:sheetData/s:row", NS)
        ]
        assert rows == ["1", "3", "5"]
        wb = _load_wb(out)
        assert wb["Notes"]["A3"].value == "three"
        assert wb["Notes"]["A5"].value == "five"

    def test_new_cells_inserted_sorted_within_row(self):
        raw = open_raw(_build_workbook(), extension="xlsx")
        sales = raw.sheets["Sales"]
        sales.set_cell("D2", 1)
        sales.set_cell("C2", 2)  # lands BETWEEN B2 and D2
        out = raw.to_bytes()
        row2 = _sheet_root(out, SHEET1).findall("s:sheetData/s:row", NS)[1]
        assert [c.get("r") for c in row2] == ["A2", "B2", "C2", "D2"]
