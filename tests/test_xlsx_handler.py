"""
Comprehensive tests for the XLSX handler pipeline.

Tests cover:
1. Constants
2. XlsxConverter: ZIP validation, openpyxl opening, close
3. XlsxPreprocessor: Workbook wrapping, resource pre-extraction
4. XlsxMetadataExtractor: OOXML core properties
5. Layout detection: LayoutRange, layout_detect_range, object_detect
6. Table conversion: Markdown and HTML output
7. XlsxContentExtractor: Full text extraction, charts, images
8. XLSXHandler: Integration, pipeline
9. Chart XML parsing helpers
10. Edge cases
"""

from __future__ import annotations

import io
import os
import zipfile
from datetime import datetime
from typing import Any
from unittest.mock import MagicMock, patch, PropertyMock
from xml.etree import ElementTree as ET

import openpyxl
import pytest

# ═══════════════════════════════════════════════════════════════════════════════
# Imports under test
# ═══════════════════════════════════════════════════════════════════════════════

from contextifier_new.handlers.xlsx._constants import (
    ZIP_MAGIC,
    NS_CHART,
    NS_DRAWING_MAIN,
    OOXML_NS,
    CHART_TYPE_MAP,
    SUPPORTED_IMAGE_EXTENSIONS,
    UNSUPPORTED_IMAGE_EXTENSIONS,
    MAX_SCAN_ROWS,
    MAX_SCAN_COLS,
)
from contextifier_new.handlers.xlsx.converter import (
    XlsxConverter,
    XlsxConvertedData,
)
from contextifier_new.handlers.xlsx.preprocessor import (
    XlsxPreprocessor,
    _parse_chart_xml,
    _extract_images_from_zip,
)
from contextifier_new.handlers.xlsx.metadata_extractor import XlsxMetadataExtractor
from contextifier_new.handlers.xlsx._layout import (
    LayoutRange,
    layout_detect_range,
    object_detect,
)
from contextifier_new.handlers.xlsx._table import (
    convert_region_to_table,
    convert_region_to_markdown,
    convert_region_to_html,
    convert_sheet_to_text,
)
from contextifier_new.handlers.xlsx.content_extractor import XlsxContentExtractor
from contextifier_new.handlers.xlsx.handler import XLSXHandler
from contextifier_new.types import (
    DocumentMetadata,
    ExtractionResult,
    FileContext,
    PreprocessedData,
    TableData,
    ChartData,
)
from contextifier_new.config import ProcessingConfig
from contextifier_new.errors import ConversionError, PreprocessingError


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _make_ctx(data: bytes, ext: str = "xlsx", name: str = "test.xlsx") -> FileContext:
    """Create a minimal FileContext dict."""
    return {
        "file_data": data,
        "file_extension": ext,
        "file_name": name,
        "file_path": f"/tmp/{name}",
        "file_category": "spreadsheet",
        "file_stream": io.BytesIO(data),
        "file_size": len(data),
    }


def _make_minimal_xlsx() -> bytes:
    """Create a minimal XLSX file in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header1"
    ws["B1"] = "Header2"
    ws["A2"] = "Value1"
    ws["B2"] = "Value2"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_xlsx_with_data(sheets: dict) -> bytes:
    """
    Create an XLSX file with specific sheet data.

    Args:
        sheets: dict of sheet_name → list of rows (each row is a list of values)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_xlsx_with_merge() -> bytes:
    """Create an XLSX file with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["A1"] = "Merged Header"
    ws.merge_cells("A1:B1")
    ws["A2"] = "Val1"
    ws["B2"] = "Val2"
    ws["A3"] = "Val3"
    ws["B3"] = "Val4"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_xlsx_with_metadata() -> bytes:
    """Create an XLSX file with metadata properties."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Test"
    wb.properties.title = "Test Title"
    wb.properties.subject = "Test Subject"
    wb.properties.creator = "Test Author"
    wb.properties.keywords = "test, xlsx"
    wb.properties.description = "Test Description"
    wb.properties.lastModifiedBy = "Modifier"
    wb.properties.category = "Category"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _make_xlsx_multi_sheet() -> bytes:
    """Create an XLSX with multiple sheets."""
    return _make_xlsx_with_data({
        "Sales": [
            ["Product", "Revenue"],
            ["Widget A", 1000],
            ["Widget B", 2000],
        ],
        "Inventory": [
            ["Item", "Qty"],
            ["Widget A", 50],
            ["Widget B", 100],
        ],
    })


def _make_chart_xml(
    chart_type: str = "barChart",
    title: str = "Sales Chart",
    categories: list = None,
    series: list = None,
) -> bytes:
    """Create a minimal OOXML chart XML."""
    ns_c = NS_CHART
    ns_a = NS_DRAWING_MAIN

    cats = categories or ["Q1", "Q2", "Q3"]
    sers = series or [{"name": "Revenue", "values": [100, 200, 300]}]

    xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<c:chartSpace xmlns:c="{ns_c}" xmlns:a="{ns_a}">
  <c:chart>
    <c:title>
      <c:tx>
        <c:rich>
          <a:p>
            <a:r>
              <a:t>{title}</a:t>
            </a:r>
          </a:p>
        </c:rich>
      </c:tx>
    </c:title>
    <c:plotArea>
      <c:{chart_type}>"""

    for s in sers:
        xml += f"""
        <c:ser>
          <c:tx><c:v>{s["name"]}</c:v></c:tx>
          <c:cat>
            <c:strCache>"""
        for i, cat in enumerate(cats):
            xml += f"""
              <c:pt idx="{i}"><c:v>{cat}</c:v></c:pt>"""
        xml += """
            </c:strCache>
          </c:cat>
          <c:val>
            <c:numCache>"""
        for i, val in enumerate(s["values"]):
            xml += f"""
              <c:pt idx="{i}"><c:v>{val}</c:v></c:pt>"""
        xml += """
            </c:numCache>
          </c:val>
        </c:ser>"""

    xml += f"""
      </c:{chart_type}>
    </c:plotArea>
  </c:chart>
</c:chartSpace>"""

    return xml.encode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Constants
# ═══════════════════════════════════════════════════════════════════════════════

class TestConstants:
    def test_zip_magic(self):
        assert ZIP_MAGIC == b"PK\x03\x04"

    def test_chart_type_map_has_entries(self):
        assert len(CHART_TYPE_MAP) == 16

    def test_chart_type_map_bar(self):
        assert CHART_TYPE_MAP["barChart"] == "Bar Chart"

    def test_supported_images(self):
        assert ".png" in SUPPORTED_IMAGE_EXTENSIONS
        assert ".jpg" in SUPPORTED_IMAGE_EXTENSIONS

    def test_unsupported_images(self):
        assert ".emf" in UNSUPPORTED_IMAGE_EXTENSIONS

    def test_scan_limits(self):
        assert MAX_SCAN_ROWS == 1000
        assert MAX_SCAN_COLS == 100

    def test_ooxml_ns_keys(self):
        assert "c" in OOXML_NS
        assert "a" in OOXML_NS
        assert "xdr" in OOXML_NS


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Converter
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsxConverter:
    def test_format_name(self):
        c = XlsxConverter()
        assert c.get_format_name() == "xlsx"

    def test_validate_valid_xlsx(self):
        c = XlsxConverter()
        data = _make_minimal_xlsx()
        ctx = _make_ctx(data)
        assert c.validate(ctx) is True

    def test_validate_empty(self):
        c = XlsxConverter()
        ctx = _make_ctx(b"")
        assert c.validate(ctx) is False

    def test_validate_non_zip(self):
        c = XlsxConverter()
        ctx = _make_ctx(b"not a zip file" * 10)
        assert c.validate(ctx) is False

    def test_validate_zip_without_content_types(self):
        """A plain ZIP without [Content_Types].xml is not valid XLSX."""
        c = XlsxConverter()
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("hello.txt", "world")
        ctx = _make_ctx(buf.getvalue())
        assert c.validate(ctx) is False

    def test_convert_success(self):
        c = XlsxConverter()
        data = _make_minimal_xlsx()
        ctx = _make_ctx(data)
        result = c.convert(ctx)
        assert isinstance(result, XlsxConvertedData)
        assert isinstance(result.workbook, openpyxl.Workbook)
        assert result.file_data == data
        result.workbook.close()

    def test_convert_empty_raises(self):
        c = XlsxConverter()
        ctx = _make_ctx(b"")
        with pytest.raises(ConversionError, match="Empty"):
            c.convert(ctx)

    def test_convert_invalid_raises(self):
        c = XlsxConverter()
        ctx = _make_ctx(b"garbage data" * 100)
        with pytest.raises(ConversionError):
            c.convert(ctx)

    def test_close_converted_data(self):
        c = XlsxConverter()
        data = _make_minimal_xlsx()
        ctx = _make_ctx(data)
        result = c.convert(ctx)
        c.close(result)  # Should not raise

    def test_close_workbook_directly(self):
        c = XlsxConverter()
        wb = openpyxl.Workbook()
        c.close(wb)  # Should not raise

    def test_close_none(self):
        c = XlsxConverter()
        c.close(None)  # Should not raise

    def test_converted_data_namedtuple(self):
        wb = MagicMock()
        cd = XlsxConvertedData(workbook=wb, file_data=b"test")
        assert cd.workbook is wb
        assert cd.file_data == b"test"


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Preprocessor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsxPreprocessor:
    def test_format_name(self):
        p = XlsxPreprocessor()
        assert p.get_format_name() == "xlsx"

    def test_preprocess_none_raises(self):
        p = XlsxPreprocessor()
        with pytest.raises(PreprocessingError, match="None"):
            p.preprocess(None)

    def test_preprocess_basic(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        cd = XlsxConvertedData(workbook=wb, file_data=data)

        p = XlsxPreprocessor()
        result = p.preprocess(cd)

        assert isinstance(result, PreprocessedData)
        assert result.content is wb
        assert result.properties["sheet_count"] == 1
        assert "Sheet1" in result.properties["sheet_names"]
        wb.close()

    def test_preprocess_multi_sheet(self):
        data = _make_xlsx_multi_sheet()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        cd = XlsxConvertedData(workbook=wb, file_data=data)

        p = XlsxPreprocessor()
        result = p.preprocess(cd)

        assert result.properties["sheet_count"] == 2
        assert "Sales" in result.properties["sheet_names"]
        assert "Inventory" in result.properties["sheet_names"]
        wb.close()

    def test_preprocess_without_file_data(self):
        """Preprocessor should work without file_data (no ZIP extraction)."""
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        # Pass workbook directly (no XlsxConvertedData)
        # This will hit the isinstance(openpyxl.Workbook) branch
        p = XlsxPreprocessor()
        result = p.preprocess(wb)

        assert result.content is wb
        assert result.resources["charts"] == []
        assert result.resources["images"] == {}
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Metadata Extractor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsxMetadataExtractor:
    def test_format_name(self):
        m = XlsxMetadataExtractor()
        assert m.get_format_name() == "xlsx"

    def test_extract_from_none(self):
        m = XlsxMetadataExtractor()
        result = m.extract(None)
        assert isinstance(result, DocumentMetadata)

    def test_extract_full_metadata(self):
        data = _make_xlsx_with_metadata()
        wb = openpyxl.load_workbook(io.BytesIO(data))

        m = XlsxMetadataExtractor()
        result = m.extract(wb)

        assert result.title == "Test Title"
        assert result.subject == "Test Subject"
        assert result.author == "Test Author"
        assert result.keywords == "test, xlsx"
        assert result.comments == "Test Description"
        assert result.last_saved_by == "Modifier"
        assert result.category == "Category"
        wb.close()

    def test_extract_from_workbook(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data))

        m = XlsxMetadataExtractor()
        result = m.extract(wb)
        assert isinstance(result, DocumentMetadata)
        assert result.page_count == 1
        wb.close()

    def test_extract_from_converted_data(self):
        data = _make_xlsx_with_metadata()
        wb = openpyxl.load_workbook(io.BytesIO(data))
        cd = XlsxConvertedData(workbook=wb, file_data=data)

        m = XlsxMetadataExtractor()
        result = m.extract(cd)
        assert result.title == "Test Title"
        wb.close()

    def test_safe_str_none(self):
        assert XlsxMetadataExtractor._safe_str(None) is None

    def test_safe_str_empty(self):
        assert XlsxMetadataExtractor._safe_str("") is None

    def test_safe_str_value(self):
        assert XlsxMetadataExtractor._safe_str("hello") == "hello"

    def test_safe_datetime_none(self):
        assert XlsxMetadataExtractor._safe_datetime(None) is None

    def test_safe_datetime_value(self):
        dt = datetime(2024, 1, 1)
        assert XlsxMetadataExtractor._safe_datetime(dt) == dt


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Layout Detection
# ═══════════════════════════════════════════════════════════════════════════════

class TestLayoutRange:
    def test_basic_properties(self):
        lr = LayoutRange(min_row=1, max_row=5, min_col=1, max_col=3)
        assert lr.rows == 5
        assert lr.cols == 3
        assert lr.cell_count == 15

    def test_contains(self):
        lr = LayoutRange(min_row=2, max_row=10, min_col=3, max_col=8)
        assert lr.contains(5, 5) is True
        assert lr.contains(1, 1) is False
        assert lr.contains(2, 3) is True  # corner
        assert lr.contains(10, 8) is True  # opposite corner

    def test_overlaps(self):
        lr1 = LayoutRange(min_row=1, max_row=5, min_col=1, max_col=5)
        lr2 = LayoutRange(min_row=3, max_row=7, min_col=3, max_col=7)
        assert lr1.overlaps(lr2) is True
        assert lr2.overlaps(lr1) is True

    def test_no_overlap(self):
        lr1 = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=2)
        lr2 = LayoutRange(min_row=5, max_row=6, min_col=5, max_col=6)
        assert lr1.overlaps(lr2) is False

    def test_is_adjacent(self):
        lr1 = LayoutRange(min_row=1, max_row=3, min_col=1, max_col=3)
        lr2 = LayoutRange(min_row=1, max_row=3, min_col=4, max_col=6)
        assert lr1.is_adjacent(lr2) is True

    def test_merge_with(self):
        lr1 = LayoutRange(min_row=1, max_row=3, min_col=1, max_col=3)
        lr2 = LayoutRange(min_row=5, max_row=7, min_col=2, max_col=4)
        merged = lr1.merge_with(lr2)
        assert merged.min_row == 1
        assert merged.max_row == 7
        assert merged.min_col == 1
        assert merged.max_col == 4


class TestLayoutDetection:
    def test_detect_range_basic(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Hello"
        ws["C3"] = "World"

        layout = layout_detect_range(ws)
        assert layout is not None
        assert layout.min_row == 2
        assert layout.min_col == 2
        assert layout.max_row == 3
        assert layout.max_col == 3
        wb.close()

    def test_detect_range_empty_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        layout = layout_detect_range(ws)
        assert layout is None
        wb.close()

    def test_detect_range_single_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Only"

        layout = layout_detect_range(ws)
        assert layout is not None
        assert layout.rows == 1
        assert layout.cols == 1
        wb.close()

    def test_object_detect_basic(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "H1"
        ws["B1"] = "H2"
        ws["A2"] = "V1"
        ws["B2"] = "V2"

        regions = object_detect(ws)
        assert len(regions) >= 1
        wb.close()

    def test_object_detect_empty_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        regions = object_detect(ws)
        assert regions == []
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Table Conversion
# ═══════════════════════════════════════════════════════════════════════════════

class TestTableConversion:
    def test_markdown_basic(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Item"
        ws["B2"] = 42

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=2)
        text = convert_region_to_markdown(ws, region)

        assert "Name" in text
        assert "Value" in text
        assert "Item" in text
        assert "42" in text
        assert "---" in text  # Separator
        assert "|" in text
        wb.close()

    def test_html_basic(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A2"] = "Data"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=1)
        html = convert_region_to_html(ws, region)

        assert "<table>" in html
        assert "<th>" in html
        assert "Header" in html
        assert "<td>" in html
        assert "Data" in html
        wb.close()

    def test_html_with_merged_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Merged"
        ws.merge_cells("A1:B1")
        ws["A2"] = "L"
        ws["B2"] = "R"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=2)
        html = convert_region_to_html(ws, region)

        assert "colspan" in html
        assert "Merged" in html
        wb.close()

    def test_convert_sheet_to_text_selects_html_for_merge(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Merged Header"
        ws.merge_cells("A1:C1")
        ws["A2"] = "a"
        ws["B2"] = "b"
        ws["C2"] = "c"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=3)
        text = convert_sheet_to_text(ws, region)

        assert "<table>" in text
        wb.close()

    def test_convert_sheet_to_text_selects_md_no_merge(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "H"
        ws["A2"] = "V"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=1)
        text = convert_sheet_to_text(ws, region)

        assert "<table>" not in text
        assert "|" in text
        wb.close()

    def test_convert_region_to_table(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "H1"
        ws["B1"] = "H2"
        ws["A2"] = "V1"
        ws["B2"] = "V2"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=2)
        table = convert_region_to_table(ws, region)

        assert table is not None
        assert isinstance(table, TableData)
        assert table.num_rows == 2
        assert table.num_cols == 2
        wb.close()

    def test_convert_region_to_table_empty(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        region = LayoutRange(min_row=1, max_row=3, min_col=1, max_col=3)
        table = convert_region_to_table(ws, region)

        assert table is None
        wb.close()

    def test_format_cell_float_integer(self):
        """Floats that are integers should format without decimals."""
        from contextifier_new.handlers.xlsx._table import _format_cell_value
        assert _format_cell_value(42.0) == "42"
        assert _format_cell_value(3.14) == "3.14"
        assert _format_cell_value(None) == ""
        assert _format_cell_value(True) == "TRUE"
        assert _format_cell_value(False) == "FALSE"


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Content Extractor
# ═══════════════════════════════════════════════════════════════════════════════

class TestXlsxContentExtractor:
    def test_format_name(self):
        ext = XlsxContentExtractor()
        assert ext.get_format_name() == "xlsx"

    def test_extract_text_basic(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)

        assert "[Sheet: Sheet1]" in result
        assert "Header1" in result
        assert "Value1" in result
        wb.close()

    def test_extract_text_multi_sheet(self):
        data = _make_xlsx_multi_sheet()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)

        assert "[Sheet: Sales]" in result
        assert "[Sheet: Inventory]" in result
        assert "Widget A" in result
        wb.close()

    def test_extract_text_empty_workbook(self):
        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={},
        )
        result = ext.extract_text(ppd)
        assert result == ""

    def test_extract_text_with_tag_service(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        tag_svc = MagicMock()
        tag_svc.make_sheet_tag.return_value = "[SHEET: Sheet1]"

        ext = XlsxContentExtractor(tag_service=tag_svc)
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)
        assert "[SHEET: Sheet1]" in result
        wb.close()

    def test_extract_text_with_textboxes(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={
                "charts": [],
                "images": {},
                "textboxes": {"Sheet1": ["Textbox content here"]},
            },
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)
        assert "Textbox content here" in result
        wb.close()

    def test_extract_tables(self):
        data = _make_minimal_xlsx()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        tables = ext.extract_tables(ppd)
        assert len(tables) >= 1
        assert all(isinstance(t, TableData) for t in tables)
        wb.close()

    def test_extract_images_no_service(self):
        ext = XlsxContentExtractor(image_service=None)
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="utf-8",
            resources={"charts": [], "images": {"xl/media/img.png": b"\x89PNG"}, "textboxes": {}},
            properties={},
        )
        result = ext.extract_images(ppd)
        assert result == []

    def test_extract_images_with_service(self):
        img_svc = MagicMock()
        img_svc.save_and_tag.return_value = "[IMG:excel_img]"

        ext = XlsxContentExtractor(image_service=img_svc)
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="utf-8",
            resources={
                "charts": [],
                "images": {"xl/media/image1.png": b"\x89PNG" + b"\x00" * 30},
                "textboxes": {},
            },
            properties={},
        )
        result = ext.extract_images(ppd)
        assert len(result) == 1
        img_svc.save_and_tag.assert_called_once()

    def test_extract_images_dedup(self):
        img_svc = MagicMock()
        img_svc.save_and_tag.return_value = "[IMG:excel_img]"

        same_data = b"\x89PNG" + b"\x00" * 30
        ext = XlsxContentExtractor(image_service=img_svc)
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="utf-8",
            resources={
                "charts": [],
                "images": {
                    "xl/media/image1.png": same_data,
                    "xl/media/image2.png": same_data,
                },
                "textboxes": {},
            },
            properties={},
        )
        result = ext.extract_images(ppd)
        assert len(result) == 1

    def test_extract_charts(self):
        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=None,
            raw_content=None,
            encoding="utf-8",
            resources={
                "charts": [
                    {
                        "chart_type": "Bar Chart",
                        "title": "Sales",
                        "categories": ["Q1", "Q2"],
                        "series": [{"name": "Revenue", "values": [100, 200]}],
                    }
                ],
                "images": {},
                "textboxes": {},
            },
            properties={},
        )
        charts = ext.extract_charts(ppd)
        assert len(charts) == 1
        assert isinstance(charts[0], ChartData)
        assert charts[0].chart_type == "Bar Chart"
        assert charts[0].title == "Sales"

    def test_make_sheet_tag_no_service(self):
        ext = XlsxContentExtractor(tag_service=None)
        tag = ext._make_sheet_tag("MySheet")
        assert tag == "[Sheet: MySheet]"

    def test_format_chart_fallback(self):
        ext = XlsxContentExtractor()
        chart = {"chart_type": "Line Chart", "title": "Trend", "series": []}
        result = ext._format_chart(chart)
        assert "[Chart: Line Chart - Trend]" in result


# ═══════════════════════════════════════════════════════════════════════════════
# 8. XLSXHandler
# ═══════════════════════════════════════════════════════════════════════════════

class TestXLSXHandler:
    def test_supported_extensions(self):
        handler = XLSXHandler(ProcessingConfig())
        assert handler.supported_extensions == frozenset({"xlsx"})

    def test_handler_name(self):
        handler = XLSXHandler(ProcessingConfig())
        assert handler.handler_name == "XLSX Handler"

    def test_creates_converter(self):
        handler = XLSXHandler(ProcessingConfig())
        assert isinstance(handler.create_converter(), XlsxConverter)

    def test_creates_preprocessor(self):
        handler = XLSXHandler(ProcessingConfig())
        assert isinstance(handler.create_preprocessor(), XlsxPreprocessor)

    def test_creates_metadata_extractor(self):
        handler = XLSXHandler(ProcessingConfig())
        assert isinstance(handler.create_metadata_extractor(), XlsxMetadataExtractor)

    def test_creates_content_extractor(self):
        handler = XLSXHandler(ProcessingConfig())
        assert isinstance(handler.create_content_extractor(), XlsxContentExtractor)


# ═══════════════════════════════════════════════════════════════════════════════
# 9. Chart XML Parsing
# ═══════════════════════════════════════════════════════════════════════════════

class TestChartParsing:
    def test_parse_bar_chart(self):
        xml = _make_chart_xml("barChart", "Revenue")
        chart = _parse_chart_xml(xml)
        assert chart is not None
        assert chart["chart_type"] == "Bar Chart"
        assert chart["title"] == "Revenue"

    def test_parse_line_chart(self):
        xml = _make_chart_xml("lineChart", "Trend")
        chart = _parse_chart_xml(xml)
        assert chart is not None
        assert chart["chart_type"] == "Line Chart"

    def test_parse_pie_chart(self):
        xml = _make_chart_xml("pieChart", "Distribution")
        chart = _parse_chart_xml(xml)
        assert chart is not None
        assert chart["chart_type"] == "Pie Chart"

    def test_parse_with_series(self):
        xml = _make_chart_xml(
            "barChart",
            "Sales",
            categories=["A", "B"],
            series=[{"name": "S1", "values": [10, 20]}],
        )
        chart = _parse_chart_xml(xml)
        assert len(chart["series"]) == 1
        assert chart["series"][0]["name"] == "S1"

    def test_parse_categories(self):
        xml = _make_chart_xml(
            "barChart",
            "Sales",
            categories=["Jan", "Feb", "Mar"],
        )
        chart = _parse_chart_xml(xml)
        assert chart["categories"] == ["Jan", "Feb", "Mar"]

    def test_parse_invalid_xml(self):
        result = _parse_chart_xml(b"not xml at all")
        assert result is None

    def test_parse_xml_without_plot_area(self):
        xml = f"""<?xml version="1.0"?>
<c:chartSpace xmlns:c="{NS_CHART}">
  <c:chart>
    <c:title/>
  </c:chart>
</c:chartSpace>""".encode()
        result = _parse_chart_xml(xml)
        assert result is None

    def test_parse_xml_with_bom(self):
        xml = b"\xef\xbb\xbf" + _make_chart_xml("barChart", "BOM Test")
        chart = _parse_chart_xml(xml)
        assert chart is not None
        assert chart["title"] == "BOM Test"


# ═══════════════════════════════════════════════════════════════════════════════
# 10. Full Pipeline & Edge Cases
# ═══════════════════════════════════════════════════════════════════════════════

class TestFullPipeline:
    def test_full_pipeline_basic(self):
        data = _make_minimal_xlsx()
        handler = XLSXHandler(ProcessingConfig())
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert isinstance(result, ExtractionResult)
        assert "Header1" in result.text
        assert "Value1" in result.text

    def test_full_pipeline_multi_sheet(self):
        data = _make_xlsx_multi_sheet()
        handler = XLSXHandler(ProcessingConfig())
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert "Sales" in result.text
        assert "Inventory" in result.text
        assert "Widget A" in result.text

    def test_full_pipeline_with_metadata(self):
        data = _make_xlsx_with_metadata()
        handler = XLSXHandler(ProcessingConfig())
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert result.metadata is not None
        assert result.metadata.title == "Test Title"
        assert result.metadata.author == "Test Author"

    def test_full_pipeline_merged_cells(self):
        data = _make_xlsx_with_merge()
        handler = XLSXHandler(ProcessingConfig())
        ctx = _make_ctx(data)
        result = handler.process(ctx)

        assert "Merged Header" in result.text


class TestEdgeCases:
    def test_empty_sheet(self):
        data = _make_xlsx_with_data({"Empty": []})
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)
        # Should have sheet tag but no data
        assert "[Sheet: Empty]" in result
        wb.close()

    def test_single_cell_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Single"
        ws["A1"] = "Lonely"

        ext = XlsxContentExtractor()
        ppd = PreprocessedData(
            content=wb,
            raw_content=wb,
            encoding="utf-8",
            resources={"charts": [], "images": {}, "textboxes": {}},
            properties={"sheet_names": wb.sheetnames},
        )
        result = ext.extract_text(ppd)
        assert "Lonely" in result
        wb.close()

    def test_image_extraction_from_zip(self):
        """Test ZIP-level image extraction."""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("[Content_Types].xml", "<Types/>")
            zf.writestr("xl/media/image1.png", b"\x89PNG" + b"\x00" * 20)
            zf.writestr("xl/media/image2.jpg", b"\xFF\xD8" + b"\x00" * 20)
            zf.writestr("xl/media/image3.emf", b"\x01" * 20)  # Should be skipped

        images = _extract_images_from_zip(buf.getvalue())
        assert "xl/media/image1.png" in images
        assert "xl/media/image2.jpg" in images
        assert "xl/media/image3.emf" not in images

    def test_converter_rejects_non_xlsx(self):
        c = XlsxConverter()
        ctx = _make_ctx(b"This is not an xlsx file at all")
        with pytest.raises(ConversionError):
            c.convert(ctx)

    def test_markdown_pipe_escaping(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "A|B"
        ws["A2"] = "C"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=1)
        text = convert_region_to_markdown(ws, region)
        assert "\\|" in text
        wb.close()

    def test_html_escaping(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "<script>alert('xss')</script>"
        ws["A2"] = "Safe"

        region = LayoutRange(min_row=1, max_row=2, min_col=1, max_col=1)
        html = convert_region_to_html(ws, region)
        assert "&lt;script&gt;" in html
        assert "<script>" not in html
        wb.close()

    def test_xlsx_converted_data_fields(self):
        wb = MagicMock()
        cd = XlsxConvertedData(workbook=wb, file_data=b"\x00")
        assert cd[0] is wb
        assert cd[1] == b"\x00"
